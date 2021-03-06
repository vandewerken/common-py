import os
import openpyxl
from glob import glob
import shutil

### Set the following variables for the directory of the files you're working with

processDirectory = 
col = 

### 

def copyRange(startCol, startRow, endCol, endRow, sheet):
        rangeSelected = []
        for i in range(startRow,endRow + 1,1):
                rowSelected = []
                for j in range(startCol,endCol+1,1):
                        rowSelected.append(sheet.cell(row = i, column = j).value)
                rangeSelected.append(rowSelected)
        return rangeSelected

def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving, copiedData):
        countRow = 0
        for i in range(startRow,endRow+1,1):
                countCol = 0
                for j in range(startCol,endCol+1,1):
                        sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
                        countCol += 1
                countRow += 1

for xlsxToProcess in glob(processDirectory + "*.xlsx"): 
    fileInProcess = os.path.basename(xlsxToProcess)
    wb = openpyxl.load_workbook(xlsxToProcess)
    ws = wb.worksheets[0]
    rowCount = ws.max_row

    createNewWorkbookPath = processDirectory + "new-" + fileInProcess
    tempFile = openpyxl.Workbook()
    tempFile.save(createNewWorkbookPath)

    template = openpyxl.load_workbook(createNewWorkbookPath)
    tempSheet = template.worksheets[0]

    print("We're going to copy data from: %s" % fileInProcess)

    # The following two lines can be used for any range desired within a particular worksheet
    # If you're trying to copy non-contiguous ranges, copy the two rows and specify additional ranges - the code was tested with more than one range

    selectedRange = copyRange(int(col), 2, int(col), int(rowCount), ws) # start col, start row, end col, end row, worksheet
    pastingRange = pasteRange(1, 2, 1, int(rowCount), tempSheet, selectedRange) # start col, start row, end col, end row, new wb, ws data

    template.save(createNewWorkbookPath)
    print("\tRange copied and pasted!\n")

# If you'd like to archive the old source files, include this: 
for xlsxMoveToArchive in glob(processDirectory + "*.xlsx"):
    if "new-" not in xlsxMoveToArchive:
        shutil.move(xlsxMoveToArchive, processDirectory + "Archive/")

print("\n\n###\n" +
    "### The mod script was a success\n" + 
    "###")